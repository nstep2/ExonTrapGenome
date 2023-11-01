



outdir = exp_output_path.HEK293_recovery +'HEK293_transcript_RPKM.pdf'
pdf_plots = PdfPages(outdir)





#HEK293_exp_file = '/mnt/hgfs/main_ssd/Linux_2TB/HEK293/GSE235387_HEK293.xlsx'


HEK293_exp_file =  exp_output_path.HEK293_exp_file



from openpyxl import load_workbook

AK=36       #'HEK293-WT-S20190326-S20190326'
AL=37       #'HEK293-WT-S20200115-S20200115'

# Load the workbook
workbook = load_workbook(filename= HEK293_exp_file)

# Select the first sheet
worksheet = workbook.active

exp_1_hek293 = dict()
exp_2_hek293 = dict()
# Loop through all rows
for ii, row in enumerate(worksheet.iter_rows(values_only=True)):
    if ii == 1:
        continue
    #if ii == 2:
    #    print(row)
    key = row[0]
    exp_1_hek293[key] = row[AK]
    exp_2_hek293[key] = row[AL]













    
    
    
chr_list = ['%d'%d for d in range(1,24)]    
chr_list += ['X', 'Y']



HEK293_exp_means_dict = exp_1_hek293

    
probs=list()
accepted_gene_id_list = list()
for ii, gene_id in enumerate(HEK293_exp_means_dict):
    
    gene_id_base = gene_id  #.split('.')[0]
    
    try:
        #gene_stats = ensembl_data.gene_ids_of_gene_name(gene_id_base)
        gene_stats = ensembl_data.gene_by_id(gene_id)

    except ValueError as e:
        1
    
        probs.append([gene_id, gene_id_base])
        continue
    
    
    start  = gene_stats.start
    end    = gene_stats.end
    chrom  = gene_stats.contig
    strand = gene_stats.strand
    
    if chrom not in chr_list:
        continue
    chrom = 'chr' + str(chrom)
    
    intervals = gencode_pc_exon_IT[chrom][strand][start:end]    
    if len(intervals) >= 1:
        accepted_gene_id_list.append(gene_id)
    
    for interval in intervals:
        exon_list = interval[2]

        









gene_and_exp_list = [ [key, HEK293_exp_means_dict[key]] for key in accepted_gene_id_list] # if key in accepted_gene_id_list]
gene_and_exp_list = sorted(gene_and_exp_list, key=lambda x: x[1])
gene_and_nonzero_exp_list = [entry for entry in gene_and_exp_list if entry[1] > 0]
#sort genes by expression, highest expression at bottom
gene_and_nonzero_exp_list = sorted(gene_and_nonzero_exp_list, key=lambda x: x[1])






twenty_percent = int( len(accepted_gene_id_list)/5 )

top_20_pc_exons_list = list()
for entry in gene_and_nonzero_exp_list[-1*twenty_percent:]:
    gene_id = entry[0]
    gene_id_base = gene_id.split('.')[0]
    gene_stats = ensembl_data.locus_of_gene_id(gene_id_base)
    
    start  = gene_stats.start
    end    = gene_stats.end
    chrom  = gene_stats.contig
    strand = gene_stats.strand
    
    if chrom == 'MT':
        continue
    chrom = 'chr' + str(chrom)

    intervals = gencode_pc_exon_IT[chrom][strand][start:end]   
    
    
    for interval in intervals:
        top_20_pc_exons_list.append(interval[2])


top_20_pc_exons_list = list(set(top_20_pc_exons_list))

top_20_pc_exons_list = el.exon_id_intersection(top_20_pc_exons_list, pc_middle_exon_id_list)


top_20_pc_ET_exons_list = el.exon_id_intersection(top_20_pc_exons_list, aggregate_exon_dict.keys())









bot_20_pc_exons_list = list()
for entry in gene_and_nonzero_exp_list[:-1*twenty_percent]:
    gene_id = entry[0]
    gene_id_base = gene_id.split('.')[0]
    gene_stats = ensembl_data.locus_of_gene_id(gene_id_base)
    
    start  = gene_stats.start
    end    = gene_stats.end
    chrom  = gene_stats.contig
    strand = gene_stats.strand
    
    if chrom == 'MT':
        continue
    chrom = 'chr' + str(chrom)

    intervals = gencode_pc_exon_IT[chrom][strand][start:end]   
    
    
    for interval in intervals:
        bot_20_pc_exons_list.append(interval[2])

bot_20_pc_exons_list = list(set(bot_20_pc_exons_list))

bot_20_pc_exons_list = el.exon_id_intersection(bot_20_pc_exons_list, pc_middle_exon_id_list)


bot_20_pc_ET_exons_list = el.exon_id_intersection(bot_20_pc_exons_list, aggregate_exon_dict.keys())

len(bot_20_pc_ET_exons_list)/len(bot_20_pc_exons_list)









middle_20_pc_exons_list = list()
for entry in gene_and_nonzero_exp_list[2*twenty_percent:3*twenty_percent]:
    gene_id = entry[0]
    gene_id_base = gene_id.split('.')[0]
    gene_stats = ensembl_data.locus_of_gene_id(gene_id_base)
    
    start  = gene_stats.start
    end    = gene_stats.end
    chrom  = gene_stats.contig
    strand = gene_stats.strand
    
    if chrom == 'MT':
        continue
    chrom = 'chr' + str(chrom)

    intervals = gencode_pc_exon_IT[chrom][strand][start:end]   
    
    
    for interval in intervals:
        middle_20_pc_exons_list.append(interval[2])

middle_20_pc_exons_list = list(set(middle_20_pc_exons_list))

middle_20_pc_exons_list = el.exon_id_intersection(middle_20_pc_exons_list, pc_middle_exon_id_list)


middle_20_pc_ET_exons_list = el.exon_id_intersection(middle_20_pc_exons_list, aggregate_exon_dict.keys())










#housekeeping_input_file='/home/dude/work/HEK293_exp/housekeeping/gkaa609_supplemental_files/Supplementary_Table1.txt'


housekeeping_input_file = exp_output_path.housekeeping_list



housekeeping_gene_name_list = list()
with open(housekeeping_input_file) as f:
    for line in f:
        housekeeping_gene_name_list.append(line.strip())


HK_pc_exons_list = list()
orphan_gene_id_list = list()
for entry in housekeeping_gene_name_list:
    
    gene_name = entry
    try:
        gene_id_list = ensembl_data.gene_ids_of_gene_name(gene_name)                      
    except:
        orphan_gene_id_list.append(entry)
        continue
    
    
    for gene_id in gene_id_list:
        
        gene_id_base = gene_id.split('.')[0]
        gene_stats = ensembl_data.locus_of_gene_id(gene_id_base)
        
        start  = gene_stats.start
        end    = gene_stats.end
        chrom  = gene_stats.contig
        strand = gene_stats.strand
        
        if chrom == 'MT':
            continue
        chrom = 'chr' + str(chrom)
    
        intervals = gencode_pc_exon_IT[chrom][strand][start:end]   
        
        
        for interval in intervals:
            HK_pc_exons_list.append(interval[2])
len(HK_pc_exons_list)


HK_pc_exons_list = list(set(HK_pc_exons_list))
HK_pc_exons_list = el.exon_id_intersection(HK_pc_exons_list, pc_middle_exon_id_list)
HK_pc_ET_exons_list = el.exon_id_intersection(HK_pc_exons_list, aggregate_exon_dict.keys())

len(HK_pc_ET_exons_list)/len(HK_pc_exons_list)







top_exon_counts = el.get_exon_dict_counts(top_20_pc_ET_exons_list , aggregate_exon_dict)
middle_exon_counts = el.get_exon_dict_counts(middle_20_pc_ET_exons_list , aggregate_exon_dict)
bot_exon_counts = el.get_exon_dict_counts(bot_20_pc_ET_exons_list , aggregate_exon_dict)






no_HK_top_exons_count = el.get_exon_dict_counts(set(top_20_pc_ET_exons_list).difference(HK_pc_ET_exons_list) , aggregate_exon_dict)

HK_exons_count = el.get_exon_dict_counts(set(HK_pc_ET_exons_list) , aggregate_exon_dict)
HK_top_exons_count = el.get_exon_dict_counts(set(top_20_pc_ET_exons_list).intersection(HK_pc_ET_exons_list) , aggregate_exon_dict)
all_others = el.get_exon_dict_counts( set(top_20_pc_ET_exons_list+middle_20_pc_ET_exons_list+bot_20_pc_ET_exons_list).difference(HK_pc_ET_exons_list).difference(top_20_pc_ET_exons_list), aggregate_exon_dict )






fig = plt.figure()
plt.hist([ np.log10(HK_exons_count), np.log10(no_HK_top_exons_count), np.log10(all_others)], density=True)
plt.yscale('log')
plt.title('Exon read counts of HEK293 genes with found exon trapping exons')
plt.legend(['House Keeping Genes','top 20% genes','All other genes'])
plt.xticks([2,3,4,5],['100','1,000','10,000','100,000'])
plt.yticks([1,0.1,0.01],['1','0.1','0.01'])
plt.ylabel('exon count density')
plt.xlabel('exon sequencing read counts')
plt.tight_layout()
pdf_plots.savefig(fig)







with open(exp_output_path.HEK293_exp_file + 'S3A.txt', 'w') as f:
    #f.write('{:}\t{:}\t{:}\n'.format('Housekeeping', 'Top_20','All_others'))
    
    f.write('{:}\t{:}\n'.format('category', 'log10_read_counts'))
    
    
    f.write('housekeeping')
    for ii, val in enumerate(np.log10(HK_exons_count)):
        '\t{:}'.format(val)
    f.write('\n')
    
    f.write('Top_20')
    for ii, val in enumerate(np.log10(no_HK_top_exons_count)):
        '\t{:}'.format(val)
    f.write('\n')
    
    f.write('All_others')
    for ii, val in enumerate(np.log10(all_others)):
        '\t{:}'.format(val)
    f.write('\n')
    





pdf_plots.close()





























    
    
    
    